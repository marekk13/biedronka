import os
import io
import re
import sys
from datetime import datetime
from collections.abc import KeysView

from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
from pdf2image import convert_from_path
import pytesseract
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build, Resource
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'


class LocalFileHandler:
    def __init__(self, file_storage_path: str = 'PATH_TO_FOLDER_FOR_FILE_STORAGE') -> None:
        self.file_storage_path = file_storage_path

    def get_files(self, path, file_format) -> list[str]:
        return [os.path.join(root, file_name) for root, _, files in os.walk(path) for file_name in files if
                file_name.endswith(file_format)]

    def get_data_from_existing_txt(self) -> list[str]:
        text_all_files = []
        for file_path in self.get_files(self.file_storage_path, '.txt'):
            with open(file_path, 'r') as f:
                text_all_files.append(f.read())
        return text_all_files

    def delete_pdf_files(self) -> None:
        for pdf_file in self.get_files(self.file_storage_path, '.pdf'):
            try:
                os.remove(pdf_file)
            except OSError as e:
                print(
                    f"Error occurred while deleting PDF file {pdf_file}. Error text: {e.text}")

    def delete_txt_files(self) -> None:
        for file in self.get_files(self.file_storage_path, '.txt'):
            try:
                os.remove(file)
            except OSError as e:
                print(
                    f"Error occurred while deleting txt file {file}. Error text: {e.text}")

    def delete_local_files(self) -> None:
        self.delete_txt_files()
        self.delete_pdf_files()

    def create_txt(self, name: str, contents: str) -> None:
        with open(f'PATH_TO_FOLDER_FOR_FILE_STORAGE/{name}.txt', 'w') as f:
            f.write(contents)

    def find_files_missing_locally(self, google_drive_files: list[dict], path: str) -> list[dict[str, str]]:
        return [item for item in google_drive_files if not self.is_in_local_dir(item, path)]

    def is_in_local_dir(self, google_drive_file: dict[str, str], path: str) -> bool:
        return any(google_drive_file['name'].replace('.pdf', '.txt') in files for _, _, files in os.walk(path))


class GoogleDrive(LocalFileHandler):
    def __init__(self,
                 credentials_path: str = 'credentials.json',
                 token_path: str = 'token.json',
                 ) -> None:
        super().__init__()
        self.SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
                       'https://www.googleapis.com/auth/drive']
        self.credentials_path = credentials_path
        self.token_path = token_path

    def login(self) -> Credentials:
        try:
            creds = None
            if os.path.exists(self.token_path):
                creds = Credentials.from_authorized_user_file(
                    self.token_path, self.SCOPES)
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    flow = InstalledAppFlow.from_client_secrets_file(
                        self.credentials_path, self.SCOPES)
                    creds = flow.run_local_server(port=0)
                with open(self.token_path, 'w') as token:
                    token.write(creds.to_json())
        except HttpError as e:
            sys.exit(
                f'Error occurred while logging to Google Drive. Error text: {e.text}')
        else:
            return creds

    def download(self, creds: Credentials, latest_grocery_date_excel: datetime) -> list[str]:
        service = build('drive', 'v3', credentials=creds)
        biedronka_pdfs = self.search_valid_files(service)
        items_to_download = self.validate_to_download(
            biedronka_pdfs, latest_grocery_date_excel)
        file_paths = self.download_files(service, items_to_download)
        return file_paths

    def search_valid_files(self, service: Resource) -> list[dict]:
        query = "mimeType='application/pdf' and '1-47XS5ceGwfbOdFbZI-4mmGW2UQhqJ4S' in parents"
        results = service.files().list(
            q=query, fields="nextPageToken, files(id, name)").execute()
        items = results.get('files', [])
        return items

    def download_files(self, service: Resource, items: list[dict]) -> list[str]:
        paths = []
        for item in items:
            request = service.files().get_media(fileId=item['id'])
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fd=fh, request=request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            fh.seek(0)
            path = os.path.join(self.file_storage_path, item['name'])
            with open(path, 'wb') as f:
                f.write(fh.read())
            paths.append(path)
        return paths

    def validate_to_download(self, biedronka_pdfs: list[dict], latest_grocery_date_excel: datetime) -> list[
            dict[str, str]]:
        items = self.compare_dates(biedronka_pdfs, latest_grocery_date_excel)
        return self.find_files_missing_locally(items, self.file_storage_path)

    def compare_dates(self, items: list[dict[str, str]], latest_grocery_date_excel: datetime) -> list[dict[str, str]]:
        groceries_after_threshold = []
        for item in items:
            date_str = '20' + item['name'][:6]
            file_grocery_date = datetime.strptime(date_str, "%Y%m%d")
            if file_grocery_date > latest_grocery_date_excel:
                groceries_after_threshold.append(item)
        return groceries_after_threshold


class PDFProcessor(LocalFileHandler):
    def __init__(self, poppler_path: str = r"C:\Program Files\Release-23.08.0-0\poppler-23.08.0\Library\bin") -> None:
        super().__init__()
        self.poppler_path = poppler_path

    def get_text_from_pdf(self, pdf_file_paths: list[str]) -> list[str]:
        text_all_files = []
        for pdf_file in pdf_file_paths:
            images = convert_from_path(
                pdf_file, poppler_path=self.poppler_path)
            file_txt = ""
            for _, img in enumerate(images):
                file_txt += pytesseract.image_to_string(img, config='--psm 6')
            text_all_files.append(file_txt)
        return text_all_files

    def create_txt(self, text_all_files: list[str], pdf_file_paths: list[str]) -> None:
        for pdf_file_path, file_text in zip(pdf_file_paths, text_all_files):
            name = pdf_file_path.split('\\')[-1].split('.')[0]
            super().create_txt(name, file_text)


class DataParser:
    DATE_PATTERN = r'((?:0[1-9]|[12][0-9]|3[01])[/\-.](?:0[1-9]|1[0-2])[/\-.](?:19|20)\d{2}) (?:0\d|1\d|2[0-3]):[0-5][0-9]'
    RECORD_PATTERN = (
        r"(?:(.+) [A-Z] (?:[1-9][0-9]|[0-9])\.[0-9]{3} [xX]?[xX]? ?(?:[1-9][0-9][0-9]|[1-9][0-9]|[0-9]|\S+)(?:[,\.][0-9]{2})? (?:[1-9][0-9][0-9]|[1-9][0-9]|[0-9]),[0-9]{2}\W+"
        r"(?:.+Strona.+\W+)?Rabat [-~“]?(?:[1-9][0-9][0-9]|[1-9][0-9]|[0-9])[,\.][0-9]{2}\W+"
        r"([1-9][0-9][0-9],[0-9]{2}|[1-9][0-9],[0-9]{2}|[0-9][,\.][0-9]{2}))|"

        r"(?:(.+) [A-Z] (?:[1-9][0-9]|[0-9])\.[0-9]{3} [xX]?[xX]? ?(?:[1-9][0-9][0-9]|[1-9][0-9]|[0-9]|\S+)(?:[,\.][0-9]{2})? ((?:[1-9][0-9][0-9]|[1-9][0-9]|[0-9])[,\.][0-9]{2}))"
    )

    def __init__(self) -> None:
        pass

    def parse_data_from_pdf(self, documents: list[str]) -> list[tuple[str, list[list[str, str]]]]:
        all_doc_data = []
        date_pattern = re.compile(DataParser.DATE_PATTERN)
        record_pattern = re.compile(DataParser.RECORD_PATTERN)
        for doc in documents:
            date = date_pattern.search(doc).group(1)
            found = record_pattern.finditer(doc)
            doc_records = []
            for match in found:
                non_none_groups = [match.group(i) for i in range(
                    1, 5) if match.group(i) is not None]
                non_none_groups[1] = non_none_groups[1].replace(',', '.')
                doc_records.append(non_none_groups)
            all_doc_data.append((date, doc_records))
        return all_doc_data

    def assign_to_months(self, data: list[tuple[str, list[list[str, str]]]]) -> dict[
            str, list[tuple[str, list[list[str, str]]]]]:
        dates_d = {}
        for doc in data:
            month_year = doc[0].split(' ')[0].split('.', 1)[1]
            shoppings_in_month = dates_d.setdefault(month_year, [])
            shoppings_in_month.append(doc)
        return dates_d

    def parse_data(self, docs: list[str]) -> dict[str, list[tuple[str, list[list[str, str]]]]]:
        data = self.parse_data_from_pdf(docs)
        data_by_month = self.assign_to_months(data)
        return data_by_month


class ExcelDataDumper:
    CURRENCY_FORMAT = '_-* #,##0.00\ [$zł-409]_-;\-* #,##0.00\ [$zł-409]_-;_-* "-"??\ [$zł-409]_-;_-@_-'
    DATE_FORMAT_PATTERN = r'(?:0[1-9]|[12][0-9]|3[01])\.(?:0[1-9]|1[0-2])\.(?:19|20)\d{2}'

    def __init__(self, path_to_excel: str = 'PATH_TO_RESULTING_EXCEL_FILE') -> None:
        self.path_to_excel = path_to_excel
        try:
            self.wb = openpyxl.load_workbook(path_to_excel)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.wb.save(path_to_excel)

    def handle_sheets(self, data_keys: KeysView[str]) -> None:
        self.first_sheet_rename(data_keys)
        self.create_sheets(data_keys)

    def create_sheets(self, data_keys: KeysView[str]) -> None:
        for key in data_keys:
            if not list(filter(lambda sheetname: sheetname == key, self.wb.sheetnames)):
                self.wb.create_sheet(title=key)
        self.save_excel_workbook()

    def first_sheet_rename(self, data_keys: KeysView[str]) -> None:
        if len(self.wb.sheetnames) == 1:
            self.wb.worksheets[0].title = data_keys[0]
        self.save_excel_workbook()

    def latest_grocery_date(self) -> datetime:
        ws = self.wb.worksheets[len(self.wb.sheetnames) - 1]
        latest_grocery = datetime(1800, 1, 1, 0, 0)
        for row in ws.iter_rows(values_only=True):
            if not row[0]:
                continue
            if match := re.match(ExcelDataDumper.DATE_FORMAT_PATTERN, str(row[0])):
                grocery = datetime.strptime(match[0], "%d.%m.%Y")
                if grocery > latest_grocery:
                    latest_grocery = grocery
        return latest_grocery

    def set_number_format(self, ws: Worksheet, *columns: int) -> None:
        for col in columns:
            ws.cell(row=ws.max_row,
                    column=col).number_format = ExcelDataDumper.CURRENCY_FORMAT

    def save_excel_workbook(self) -> None:
        try:
            self.wb.save(self.path_to_excel)
        except PermissionError:
            print("Please close the file that script tries to write to,")

    def insert_data(self, data_by_month: dict[str, list[tuple[str, list[list[str, str]]]]]) -> None:
        self.handle_sheets(sorted(data_by_month.keys()))
        for month, data in sorted(data_by_month.items()):
            ws = self.wb[month]
            if ws.max_row > 4:
                ws.append([])
                ws.append([])
            for date, grocery in sorted(data):
                ws.append([date])
                ws.append(['Nazwa', 'Cena', 'NAME1', 'NAME2'])
                start_row = ws.max_row + 1
                for item_name, item_price in grocery:
                    ws.append([item_name, float(item_price), None, None])
                    self.set_number_format(ws, 2)
                end_row = ws.max_row
                ws.append(['Suma',
                           f'=SUM(B{start_row}:B{end_row})',
                           (
                               f'=SUMIF(C{start_row}:C{end_row},TRUE,B{start_row}:B{end_row})-'
                               f'SUMIFS(B{start_row}:B{end_row},D{start_row}:D{end_row},TRUE,C{start_row}:C{end_row},TRUE)/2'
                           ),
                           (
                               f'=SUMIF(D{start_row}:D{end_row},TRUE,B{start_row}:B{end_row})-'
                               f'SUMIFS(B{start_row}:B{end_row},D{start_row}:D{end_row},TRUE,C{start_row}:C{end_row},TRUE)/2'
                           )
                           ]
                          )
                self.set_number_format(ws, 2, 3, 4)
                ws.append([])
                ws.append([])
        self.save_excel_workbook()


def main():
    drive = GoogleDrive()
    creds = drive.login()

    excel = ExcelDataDumper()
    date_to_compare = excel.latest_grocery_date()

    downloaded_file_paths = drive.download(creds, date_to_compare)

    if not downloaded_file_paths:
        sys.exit('No new files to add to Excel spreadsheet.')

    pdf_processor = PDFProcessor()
    data_strings = pdf_processor.get_text_from_pdf(downloaded_file_paths)
    pdf_processor.create_txt(data_strings, downloaded_file_paths)
    pdf_processor.delete_local_files()

    data_parser = DataParser()
    parsed_data = data_parser.parse_data(data_strings)

    excel.insert_data(parsed_data)


if __name__ == "__main__":
    main()
